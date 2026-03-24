import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
import io, os, datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# ── Google Drive — legge il file Excel direttamente da Drive ──────────────────
def get_drive_service():
    """Crea il client Google Drive usando i Secrets di Streamlit."""
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds)

def load_from_drive():
    """Scarica il file Excel da Drive e restituisce i bytes."""
    try:
        service  = get_drive_service()
        file_id  = st.secrets["drive"]["file_id"]
        # Legge i metadati per sapere nome e data modifica
        meta = service.files().get(
            fileId=file_id,
            fields="name,modifiedTime"
        ).execute()
        # Scarica il contenuto
        request = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        return buf.read(), meta.get("name","portfolio.xlsx"), meta.get("modifiedTime","")
    except Exception as e:
        return None, None, str(e)

def fmt_drive_date(iso_str):
    """Formatta la data modifica Drive in italiano."""
    try:
        dt = datetime.datetime.fromisoformat(iso_str.replace("Z","+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except:
        return iso_str

st.set_page_config(page_title="Portfolio AM — Dashboard",page_icon="⚡",layout="wide",initial_sidebar_state="expanded")

st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#f4f6fa;}
[data-testid="stSidebar"]{background:#ffffff;border-right:1px solid #e0e5f0;}
[data-testid="stSidebar"] *{color:#1e2d5f!important;}
.metric-card{background:#ffffff;border-radius:12px;padding:18px 20px;margin:6px 0;
  border-left:4px solid #2e75b6;box-shadow:0 2px 8px rgba(0,0,0,0.06);}
.metric-card.green{border-left-color:#2d7a4f;}.metric-card.gold{border-left-color:#c8a227;}
.metric-card.purple{border-left-color:#7030a0;}.metric-card.blue{border-left-color:#2e75b6;}
.metric-val{font-size:26px;font-weight:700;color:#1e2d5f;margin:4px 0;}
.metric-lbl{font-size:10px;color:#6b7a99;text-transform:uppercase;letter-spacing:1px;}
.metric-sub{font-size:12px;color:#2e75b6;margin-top:2px;}
.section-title{font-size:13px;font-weight:700;color:#1e2d5f;text-transform:uppercase;
  letter-spacing:2px;border-bottom:2px solid #2e75b6;padding-bottom:6px;margin:24px 0 16px 0;}
.breadcrumb{font-size:13px;color:#6b7a99;margin-bottom:16px;padding:8px 16px;
  background:#ffffff;border-radius:8px;border:1px solid #e0e5f0;}
h1,h2,h3{color:#1e2d5f!important;}
</style>""", unsafe_allow_html=True)

@st.cache_data(ttl=300)  # cache 5 minuti — rilegge Drive automaticamente
def load_data(fc):
    wb = load_workbook(io.BytesIO(fc), data_only=True)
    wsFD = wb['Fonte_Dati']
    fd = []
    for r in range(8, wsFD.max_row+1):
        mese=wsFD.cell(r,3).value; imp=wsFD.cell(r,4).value
        en=wsFD.cell(r,5).value; magg=wsFD.cell(r,6).value
        fl=wsFD.cell(r,7).value; fn=wsFD.cell(r,9).value; tipo=wsFD.cell(r,11).value
        if imp: fd.append({'Mese':mese,'Impianto':imp,'En_Mis_kWh':en or 0,
            'En_Magg_kWh':magg or 0,'Fat_Lordo':fl or 0,'Fat_Netto':fn or 0,'Tipo':tipo or ''})
    df_fd = pd.DataFrame(fd)
    wsDB = wb['DB_Impianti']
    db = []
    for r in range(8,24):
        nome=wsDB.cell(r,3).value; tipo=wsDB.cell(r,4).value
        kwp=wsDB.cell(r,11).value; trader=wsDB.cell(r,5).value
        ragione=wsDB.cell(r,6).value; regione=wsDB.cell(r,9).value
        if nome: db.append({'Impianto':nome,'Tipo':tipo or '','kWp':kwp,
            'Trader':trader or '','RagioneSociale':ragione or '','Regione':regione or ''})
    df_db = pd.DataFrame(db)
    wsI = wb['💹 Incentivi']
    acc,con = [],[]
    for r in range(8,24):
        site=wsI.cell(r,4).value; tot=wsI.cell(r,18).value
        g=wsI.cell(r,6).value; f=wsI.cell(r,7).value; m=wsI.cell(r,8).value
        if site: acc.append({'Site':site,'Gen':g or 0,'Feb':f or 0,'Mar':m or 0,'Tot':tot or 0})
    for r in range(28,44):
        site=wsI.cell(r,4).value; tot=wsI.cell(r,18).value
        g=wsI.cell(r,6).value; f=wsI.cell(r,7).value; m=wsI.cell(r,8).value
        if site: con.append({'Site':site,'Gen':g or 0,'Feb':f or 0,'Mar':m or 0,'Tot':tot or 0})
    df_acc=pd.DataFrame(acc); df_con=pd.DataFrame(con)
    wsAP = wb['Analisi_Perdite']
    ap = []
    for r in range(41, wsAP.max_row+1):
        nome=wsAP.cell(r,3).value; tipo=wsAP.cell(r,4).value; mese=wsAP.cell(r,5).value
        et=wsAP.cell(r,9).value; er=wsAP.cell(r,10).value
        de=wsAP.cell(r,11).value; dep=wsAP.cell(r,12).value
        if nome and mese and tipo:
            ap.append({'Impianto':nome,'Tipo':tipo,'Mese':mese,'E_Teorica':et or 0,
                'E_Reale':er or 0,'Delta_MWh':de or 0,'Delta_pct':dep or 0})
    df_ap=pd.DataFrame(ap)
    return df_fd,df_db,df_acc,df_con,df_ap

def dl_excel(dfs, fname, label):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        for sn, df in dfs.items(): df.to_excel(w, sheet_name=sn[:31], index=False)
    st.download_button(label, buf.getvalue(), fname,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def kpi(lbl, val, sub="", col="blue"):
    return f'<div class="metric-card {col}"><div class="metric-lbl">{lbl}</div><div class="metric-val">{val}</div><div class="metric-sub">{sub}</div></div>'

COLORS={"Fotovoltaico":"#2d7a4f","Eolico":"#065a82"}
MESI=["Gennaio","Febbraio","Marzo"]
PBG="rgba(244,246,250,0)"; GC="#dde3ef"; FC="#3a4a65"

def cl(fig, h=360):
    fig.update_layout(plot_bgcolor=PBG,paper_bgcolor=PBG,font_color=FC,height=h,
        xaxis=dict(gridcolor=GC,linecolor=GC),yaxis=dict(gridcolor=GC,linecolor=GC),
        legend=dict(bgcolor="rgba(255,255,255,0.8)",bordercolor=GC,borderwidth=1),
        margin=dict(t=30,b=30,l=10,r=10))
    return fig

# ── Session state ─────────────────────────────────────────────────────────────
for k,v in [('lvl','portfolio'),('imp',None),('mese',None)]:
    if k not in st.session_state: st.session_state[k]=v

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    base_dir=os.path.dirname(os.path.abspath(__file__))
    lp=os.path.join(base_dir,'logo.jpg')
    if os.path.exists(lp): st.image(lp,use_container_width=True)
    else: st.markdown("### ⚡ Portfolio AM")
    st.divider()
    # ── Caricamento dati — Drive prima, upload manuale come fallback ─────────
    drive_ok = "gcp_service_account" in st.secrets and "drive" in st.secrets

    if drive_ok:
        # Pulsante aggiornamento manuale
        col_ref, col_info = st.columns([1, 2])
        with col_ref:
            force_refresh = st.button("🔄 Aggiorna", use_container_width=True)
        if force_refresh:
            st.cache_data.clear()
            st.rerun()

        # Carica da Drive
        with st.spinner("Lettura da Google Drive..."):
            fb_drive, drive_name, drive_mod = load_from_drive()
        if fb_drive:
            fb = fb_drive
            mod_fmt = fmt_drive_date(drive_mod)
            st.success(f"✅ {drive_name}")
            st.caption(f"Aggiornato il {mod_fmt}")
        else:
            st.warning("⚠ Drive non raggiungibile")
            with st.expander("🔍 Dettaglio errore"):
                st.code(str(drive_mod))
                try:
                    fid = st.secrets["drive"]["file_id"]
                    st.caption(f"File ID usato: ...{fid[-12:]}")
                except:
                    st.caption("File ID: non trovato nei Secrets")
            fb = None
    else:
        st.warning("⚠ Secrets Google non configurati — usa upload manuale")
        fb = None

    # Upload manuale (fallback o aggiornamento forzato)
    st.markdown("**Carica file manualmente**")
    up = st.file_uploader("Sostituisce temporaneamente Drive", type=['xlsx','xlsm'])
    if up:
        fb = up.read()
        st.success(f"✅ {up.name} (locale)")

    # Fallback finale: file di default nella cartella app
    if fb is None:
        data_dir    = os.path.join(base_dir, 'data')
        saved_path  = os.path.join(data_dir, 'ultimo_portfolio.xlsx')
        default_path= os.path.join(base_dir, 'portfolio_integrato_Q1_2026.xlsx')
        if os.path.exists(saved_path):
            with open(saved_path,'rb') as f: fb = f.read()
            st.info("📊 Ultimo file salvato localmente")
        elif os.path.exists(default_path):
            with open(default_path,'rb') as f: fb = f.read()
            st.info("📊 Dati Q1 2026 (default)")
        else:
            st.error("⚠ Nessun file disponibile.")
            st.stop()

    # Salva sempre una copia locale come backup
    if fb:
        data_dir = os.path.join(base_dir, 'data')
        os.makedirs(data_dir, exist_ok=True)
        with open(os.path.join(data_dir,'ultimo_portfolio.xlsx'),'wb') as f:
            f.write(fb)
    st.divider()
    st.markdown("**Navigazione**")
    if st.button("🏠 Portfolio — vista generale"):
        st.session_state.lvl='portfolio'; st.session_state.imp=None; st.session_state.mese=None
    st.divider()
    st.markdown("**Filtri**")
    s_tipo=st.multiselect("Tipo",["Fotovoltaico","Eolico"],default=["Fotovoltaico","Eolico"])
    s_mese=st.multiselect("Mese",MESI,default=MESI)

df_fd,df_db,df_acc,df_con,df_ap=load_data(fb)

# ════════════════════════════════════════════════════════════════════════════
# L1 — PORTFOLIO
# ════════════════════════════════════════════════════════════════════════════
if st.session_state.lvl=='portfolio':
    cl2,ct=st.columns([1,4])
    with cl2:
        if os.path.exists(lp): st.image(lp,width=160)
    with ct:
        st.markdown("# ⚡ Portfolio Rinnovabili — Dashboard Q1 2026")
        st.markdown("13 Impianti FV + 3 Eolici &nbsp;|&nbsp; Q1 2026")
    st.divider()

    dff=df_fd[df_fd['Tipo'].isin(s_tipo)&df_fd['Mese'].isin(s_mese)]
    tfn=dff['Fat_Netto'].sum(); tm=dff['En_Magg_kWh'].sum()
    eur=(tfn/(tm/1000)) if tm else 0
    ta=df_acc['Tot'].sum() if not df_acc.empty else 0
    tc=df_con['Tot'].sum() if not df_con.empty else 0

    c1,c2,c3,c4,c5,c6=st.columns(6)
    with c1: st.markdown(kpi("⚡ Energia",f"{dff['En_Mis_kWh'].sum()/1000:,.0f} MWh","Q1 2026","blue"),unsafe_allow_html=True)
    with c2: st.markdown(kpi("💰 Fatturato",f"€ {tfn:,.0f}","Netto Q1","green"),unsafe_allow_html=True)
    with c3: st.markdown(kpi("📈 EUR/MWh",f"€ {eur:,.2f}","Prezzo medio","gold"),unsafe_allow_html=True)
    with c4: st.markdown(kpi("💹 Acconto GSE",f"€ {ta:,.0f}","ADVANCE-GSE","purple"),unsafe_allow_html=True)
    with c5: st.markdown(kpi("✅ Consuntivo",f"€ {tc:,.0f}","OWNER×Tariffa","green"),unsafe_allow_html=True)
    with c6: st.markdown(kpi("🏭 Impianti","13 FV + 3 Eo.","Portfolio attivo","blue"),unsafe_allow_html=True)
    st.divider()

    tab1,tab2,tab3,tab4,tab5=st.tabs(["📊 Produzione","💰 Finanziario","💹 Incentivi","⚡ Perdite","🏭 Impianti"])

    with tab1:
        st.markdown('<div class="section-title">Produzione Mensile</div>',unsafe_allow_html=True)
        dm=dff.groupby(['Mese','Tipo'])['En_Mis_kWh'].sum().reset_index()
        dm['MWh']=dm['En_Mis_kWh']/1000
        dm['Mese']=pd.Categorical(dm['Mese'],categories=MESI,ordered=True)
        dm=dm.sort_values('Mese')
        c1,c2=st.columns([2,1])
        with c1:
            fig=px.bar(dm,x='Mese',y='MWh',color='Tipo',color_discrete_map=COLORS,barmode='stack',text_auto='.0f')
            fig.update_traces(textfont_color='white',textposition='inside')
            st.plotly_chart(cl(fig),use_container_width=True)
        with c2:
            dp=dff.groupby('Tipo')['En_Mis_kWh'].sum().reset_index()
            fig2=px.pie(dp,names='Tipo',values='En_Mis_kWh',color='Tipo',color_discrete_map=COLORS,hole=0.5)
            fig2.update_layout(plot_bgcolor=PBG,paper_bgcolor=PBG,font_color=FC,height=360,
                legend=dict(orientation='h',yanchor='bottom',y=-0.25))
            fig2.update_traces(textinfo='percent+label')
            st.plotly_chart(fig2,use_container_width=True)
        st.markdown('<div class="section-title">Clicca un impianto per il drill-down →</div>',unsafe_allow_html=True)
        dt=dff.groupby(['Impianto','Tipo'])['En_Mis_kWh'].sum().reset_index()
        dt['MWh']=dt['En_Mis_kWh']/1000
        dt=dt.sort_values('MWh',ascending=False)
        cols=st.columns(4)
        for i,(_, r) in enumerate(dt.reset_index(drop=True).iterrows()):
            with cols[i%4]:
                ic="💨" if r['Tipo']=='Eolico' else "☀"
                if st.button(f"{ic} {r['Impianto']}\n{r['MWh']:,.0f} MWh",key=f"p{i}",use_container_width=True):
                    st.session_state.lvl='impianto'; st.session_state.imp=r['Impianto']; st.rerun()
        st.divider()
        dl_excel({'Produzione_Q1':dt[['Impianto','Tipo','MWh']]},"produzione_Q1.xlsx","📥 Esporta Excel")

    with tab2:
        st.markdown('<div class="section-title">Finanziario</div>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        with c1:
            df2=dff.groupby(['Mese','Tipo'])['Fat_Netto'].sum().reset_index()
            df2['Mese']=pd.Categorical(df2['Mese'],categories=MESI,ordered=True)
            df2=df2.sort_values('Mese')
            fig=px.bar(df2,x='Mese',y='Fat_Netto',color='Tipo',color_discrete_map=COLORS,barmode='stack',text_auto=',.0f',labels={'Fat_Netto':'€'})
            fig.update_traces(textfont_color='white',textposition='inside')
            st.plotly_chart(cl(fig),use_container_width=True)
        with c2:
            de=dff.groupby('Impianto').agg(Fat_Netto=('Fat_Netto','sum'),En_Magg=('En_Magg_kWh','sum')).reset_index()
            de['EUR_MWh']=de['Fat_Netto']/(de['En_Magg']/1000)
            de=de[de['EUR_MWh']>0].sort_values('EUR_MWh',ascending=False)
            fig2=px.bar(de,x='Impianto',y='EUR_MWh',color='EUR_MWh',
                color_continuous_scale=[[0,'#d4e9ff'],[1,'#1e2d5f']],text_auto='.1f',labels={'EUR_MWh':'€/MWh'})
            fig2.update_layout(xaxis_tickangle=45,coloraxis_showscale=False)
            fig2.update_traces(textfont_color='white')
            st.plotly_chart(cl(fig2),use_container_width=True)
        dt=dff.groupby(['Impianto','Tipo']).agg(MWh=('En_Mis_kWh',lambda x:x.sum()/1000),
            FL=('Fat_Lordo','sum'),FN=('Fat_Netto','sum')).reset_index()
        dt['EUR']=dt['FN']/dt['MWh']
        dt.columns=['Impianto','Tipo','MWh','Fat.Lordo(€)','Fat.Netto(€)','€/MWh']
        st.dataframe(dt.style.format({'MWh':'{:,.1f}','Fat.Lordo(€)':'{:,.0f}','Fat.Netto(€)':'{:,.0f}','€/MWh':'{:,.2f}'}),use_container_width=True,hide_index=True)
        st.divider(); dl_excel({'Finanziario_Q1':dt},"finanziario_Q1.xlsx","📥 Esporta Excel")

    with tab3:
        st.markdown('<div class="section-title">Incentivi GSE</div>',unsafe_allow_html=True)
        delta=tc-ta
        c1,c2,c3=st.columns(3)
        with c1: st.metric("📊 Acconto",f"€ {ta:,.0f}")
        with c2: st.metric("✅ Consuntivo",f"€ {tc:,.0f}")
        with c3: st.metric("📐 Delta",f"€ {delta:+,.0f}",delta=f"{delta/ta*100:+.1f}%" if ta else "—")
        if not df_acc.empty and not df_con.empty:
            c1,c2=st.columns(2)
            with c1:
                am=df_acc.melt(id_vars='Site',value_vars=['Gen','Feb','Mar'],var_name='M',value_name='€')
                am=am[am['€']>0]
                fig=px.bar(am,x='Site',y='€',color='M',barmode='group',
                    color_discrete_map={'Gen':'#2e75b6','Feb':'#375623','Mar':'#7030a0'},labels={'Site':''})
                fig.update_layout(xaxis_tickangle=45)
                st.plotly_chart(cl(fig,380),use_container_width=True)
            with c2:
                dc=pd.merge(df_acc[['Site','Tot']].rename(columns={'Tot':'Acconto'}),
                    df_con[['Site','Tot']].rename(columns={'Tot':'Consuntivo'}),on='Site',how='outer').fillna(0)
                dc=dc[dc['Acconto']+dc['Consuntivo']>0]
                dm2=dc.melt(id_vars='Site',var_name='T',value_name='€')
                fig2=px.bar(dm2,x='Site',y='€',color='T',barmode='group',
                    color_discrete_map={'Acconto':'#c8a227','Consuntivo':'#2d7a4f'},labels={'Site':''})
                fig2.update_layout(xaxis_tickangle=45)
                st.plotly_chart(cl(fig2,380),use_container_width=True)
            dc['Delta']=dc['Consuntivo']-dc['Acconto']
            st.divider(); dl_excel({'Incentivi_Q1':dc},"incentivi_Q1.xlsx","📥 Esporta Excel")

    with tab4:
        st.markdown('<div class="section-title">Analisi Perdite</div>',unsafe_allow_html=True)
        if not df_ap.empty:
            apf=df_ap[df_ap['Tipo'].isin(s_tipo)&df_ap['Mese'].isin(s_mese)]
            tt=apf['E_Teorica'].sum(); tr=apf['E_Reale'].sum(); td=tt-tr
            c1,c2,c3,c4=st.columns(4)
            with c1: st.metric("E.Teorica",f"{tt:,.0f} MWh")
            with c2: st.metric("E.Reale",f"{tr:,.0f} MWh")
            with c3: st.metric("ΔE",f"{td:+,.0f} MWh")
            with c4: st.metric("ΔE%",f"{td/tt*100:+.1f}%" if tt else "—")
            ca=apf.groupby(['Impianto','Tipo']).agg(ET=('E_Teorica','sum'),ER=('E_Reale','sum')).reset_index()
            ca=ca[ca['ET']>0]
            c1,c2=st.columns(2)
            with c1:
                cm2=ca.melt(id_vars=['Impianto','Tipo'],value_vars=['ET','ER'],var_name='C',value_name='MWh')
                fig=px.bar(cm2,x='Impianto',y='MWh',color='C',barmode='group',
                    color_discrete_map={'ET':'#2e75b6','ER':'#2d7a4f'},labels={'Impianto':''})
                fig.update_layout(xaxis_tickangle=45)
                st.plotly_chart(cl(fig,400),use_container_width=True)
            with c2:
                ca['P%']=(ca['ET']-ca['ER'])/ca['ET']*100
                fig2=px.bar(ca.sort_values('P%',ascending=False),x='Impianto',y='P%',color='Tipo',
                    color_discrete_map=COLORS,text_auto='.1f',labels={'Impianto':'','P%':'Perdita %'})
                fig2.update_traces(textfont_color='white')
                fig2.update_layout(xaxis_tickangle=45)
                fig2.add_hline(y=20,line_dash="dash",line_color="#c00000",annotation_text="⚠ 20%")
                fig2.add_hline(y=8,line_dash="dash",line_color="#c8a227",annotation_text="⚡ 8%")
                st.plotly_chart(cl(fig2,400),use_container_width=True)
            ta2=apf[['Impianto','Tipo','Mese','E_Teorica','E_Reale','Delta_MWh','Delta_pct']].copy()
            ta2=ta2[ta2['E_Teorica']>0]
            ta2.columns=['Impianto','Tipo','Mese','E.Teorica(MWh)','E.Reale(MWh)','ΔE(MWh)','ΔE%']
            st.dataframe(ta2.style.format({'E.Teorica(MWh)':'{:,.1f}','E.Reale(MWh)':'{:,.1f}','ΔE(MWh)':'{:+,.1f}','ΔE%':'{:+.1%}'}),use_container_width=True,hide_index=True)
            st.divider(); dl_excel({'Perdite_Q1':ta2},"perdite_Q1.xlsx","📥 Esporta Excel")

    with tab5:
        st.markdown('<div class="section-title">Impianti — clicca per il dettaglio</div>',unsafe_allow_html=True)
        di=df_db[df_db['Tipo'].isin(s_tipo)].copy()
        dfn=df_fd[df_fd['Tipo'].isin(s_tipo)].groupby('Impianto').agg(
            En_MWh=('En_Mis_kWh',lambda x:x.sum()/1000),Fat_Netto=('Fat_Netto','sum')).reset_index()
        di=di.merge(dfn,on='Impianto',how='left')
        di['EUR_MWh']=di['Fat_Netto']/di['En_MWh']
        fig=px.scatter(di[di['En_MWh']>0],x='En_MWh',y='Fat_Netto',color='Tipo',size='En_MWh',
            color_discrete_map=COLORS,hover_name='Impianto',text='Impianto',
            labels={'En_MWh':'MWh Q1','Fat_Netto':'Fat. Netto Q1 (€)'})
        fig.update_traces(textposition='top center',textfont_size=9,textfont_color=FC)
        st.plotly_chart(cl(fig,400),use_container_width=True)
        st.markdown('<div class="section-title">Seleziona impianto per drill-down</div>',unsafe_allow_html=True)
        cols=st.columns(4)
        for i,(_,r) in enumerate(di.reset_index(drop=True).iterrows()):
            with cols[i%4]:
                ic="💨" if r['Tipo']=='Eolico' else "☀"
                fn=r.get('Fat_Netto',0) or 0
                if st.button(f"{ic} {r['Impianto']}\n€{fn:,.0f}",key=f"i5_{i}",use_container_width=True):
                    st.session_state.lvl='impianto'; st.session_state.imp=r['Impianto']; st.rerun()
        ds=di[['Impianto','Tipo','Trader','Regione','kWp','En_MWh','Fat_Netto','EUR_MWh']].copy()
        ds.columns=['Impianto','Tipo','Trader','Regione','kWp DC','MWh Q1','Fat.Netto(€)','€/MWh']
        st.dataframe(ds.style.format({'kWp DC':lambda x:f'{x:,.0f}' if pd.notna(x) and x else '—',
            'MWh Q1':'{:,.1f}','Fat.Netto(€)':'{:,.0f}','€/MWh':'{:,.2f}'}),use_container_width=True,hide_index=True)
        st.divider(); dl_excel({'Impianti':ds},"impianti_Q1.xlsx","📥 Esporta Excel")

# ════════════════════════════════════════════════════════════════════════════
# L2 — IMPIANTO
# ════════════════════════════════════════════════════════════════════════════
elif st.session_state.lvl=='impianto':
    imp=st.session_state.imp
    di=df_fd[df_fd['Impianto']==imp]
    di_info=df_db[df_db['Impianto']==imp]
    tipo=di_info['Tipo'].values[0] if not di_info.empty else ''
    ic="💨" if tipo=='Eolico' else "☀"
    trader=di_info['Trader'].values[0] if not di_info.empty else '—'
    ragione=di_info['RagioneSociale'].values[0] if not di_info.empty else '—'
    imp_color=COLORS.get(tipo,"#2e75b6")

    cb,ct=st.columns([1,5])
    with cb:
        if st.button("← Portfolio"): st.session_state.lvl='portfolio'; st.rerun()
    with ct:
        st.markdown(f"# {ic} {imp}")
        st.markdown(f"**{tipo}** &nbsp;|&nbsp; {trader} &nbsp;|&nbsp; {ragione}")
    st.markdown(f'<div class="breadcrumb">🏠 Portfolio &nbsp;›&nbsp; <b>{imp}</b></div>',unsafe_allow_html=True)
    st.divider()

    fn=di['Fat_Netto'].sum(); en=di['En_Mis_kWh'].sum()/1000
    magg=di['En_Magg_kWh'].sum()/1000; eur=fn/magg if magg else 0
    ap_imp=df_ap[df_ap['Impianto']==imp] if not df_ap.empty else pd.DataFrame()

    c1,c2,c3,c4=st.columns(4)
    with c1: st.markdown(kpi("⚡ Energia Q1",f"{en:,.1f} MWh","Misurata","blue"),unsafe_allow_html=True)
    with c2: st.markdown(kpi("💰 Fatturato",f"€ {fn:,.0f}","Netto Q1","green"),unsafe_allow_html=True)
    with c3: st.markdown(kpi("📈 EUR/MWh",f"€ {eur:,.2f}","Prezzo medio","gold"),unsafe_allow_html=True)
    with c4:
        api=ap_imp[ap_imp['E_Teorica']>0] if not ap_imp.empty else pd.DataFrame()
        if not api.empty:
            et=api['E_Teorica'].sum(); er=api['E_Reale'].sum()
            dp=(er-et)/et*100 if et else 0
            col="green" if abs(dp)<=8 else ("gold" if abs(dp)<=20 else "purple")
            st.markdown(kpi("⚡ ΔE vs Teorica",f"{dp:+.1f}%","Reale vs teorica",col),unsafe_allow_html=True)
        else: st.markdown(kpi("⚡ ΔE","—","dati non disp.","blue"),unsafe_allow_html=True)
    st.divider()

    c1,c2=st.columns(2)
    di2=di.copy(); di2['MWh']=di2['En_Mis_kWh']/1000
    di2['Mese']=pd.Categorical(di2['Mese'],categories=MESI,ordered=True); di2=di2.sort_values('Mese')
    with c1:
        st.markdown('<div class="section-title">Energia mensile (MWh)</div>',unsafe_allow_html=True)
        fig=px.bar(di2,x='Mese',y='MWh',text_auto='.1f',color_discrete_sequence=[imp_color])
        fig.update_traces(textfont_color='white',textposition='inside')
        st.plotly_chart(cl(fig,280),use_container_width=True)
    with c2:
        st.markdown('<div class="section-title">Fatturato Netto mensile (€)</div>',unsafe_allow_html=True)
        fig2=px.bar(di2,x='Mese',y='Fat_Netto',text_auto=',.0f',color_discrete_sequence=["#2d7a4f"],labels={'Fat_Netto':'€'})
        fig2.update_traces(textfont_color='white',textposition='inside')
        st.plotly_chart(cl(fig2,280),use_container_width=True)

    st.markdown('<div class="section-title">Clicca un mese per il dettaglio</div>',unsafe_allow_html=True)
    cols=st.columns(3)
    for i,(_, r) in enumerate(di2.iterrows()):
        with cols[i%3]:
            if st.button(f"**{r['Mese']}**\n{r['MWh']:,.1f} MWh | € {r['Fat_Netto']:,.0f}",
                         key=f"m{i}",use_container_width=True):
                st.session_state.lvl='mese'; st.session_state.mese=r['Mese']; st.rerun()

    if not api.empty:
        st.markdown('<div class="section-title">Analisi Perdite</div>',unsafe_allow_html=True)
        as2=api[['Mese','E_Teorica','E_Reale','Delta_MWh','Delta_pct']].copy()
        as2['Mese']=pd.Categorical(as2['Mese'],categories=MESI,ordered=True); as2=as2.sort_values('Mese')
        as2.columns=['Mese','E.Teorica(MWh)','E.Reale(MWh)','ΔE(MWh)','ΔE%']
        st.dataframe(as2.style.format({'E.Teorica(MWh)':'{:,.1f}','E.Reale(MWh)':'{:,.1f}','ΔE(MWh)':'{:+,.1f}','ΔE%':'{:+.1%}'}),use_container_width=True,hide_index=True)

    st.divider()
    de2=di2[['Mese','MWh','Fat_Lordo','Fat_Netto']].copy()
    de2.columns=['Mese','MWh','Fat.Lordo(€)','Fat.Netto(€)']
    dl_excel({imp[:31]:de2},f"{imp.replace(' ','_')}_Q1.xlsx",f"📥 Esporta {imp} Excel")

# ════════════════════════════════════════════════════════════════════════════
# L3 — MESE
# ════════════════════════════════════════════════════════════════════════════
elif st.session_state.lvl=='mese':
    imp=st.session_state.imp; mese=st.session_state.mese
    dm=df_fd[(df_fd['Impianto']==imp)&(df_fd['Mese']==mese)]
    di_info=df_db[df_db['Impianto']==imp]
    tipo=di_info['Tipo'].values[0] if not di_info.empty else ''
    ic="💨" if tipo=='Eolico' else "☀"
    imp_color=COLORS.get(tipo,"#2e75b6")

    cb1,cb2,ct=st.columns([1,1,4])
    with cb1:
        if st.button("← Portfolio"): st.session_state.lvl='portfolio'; st.rerun()
    with cb2:
        if st.button(f"← {imp[:18]}"): st.session_state.lvl='impianto'; st.rerun()
    with ct: st.markdown(f"# {ic} {imp} — {mese} 2026")
    st.markdown(f'<div class="breadcrumb">🏠 Portfolio &nbsp;›&nbsp; {imp} &nbsp;›&nbsp; <b>{mese} 2026</b></div>',unsafe_allow_html=True)
    st.divider()

    if not dm.empty:
        r=dm.iloc[0]
        em=r['En_Mis_kWh']/1000; emg=r['En_Magg_kWh']/1000
        fl=r['Fat_Lordo']; fn=r['Fat_Netto']; fee=fl-fn
        eur=fn/emg if emg else 0

        c1,c2,c3,c4,c5=st.columns(5)
        with c1: st.markdown(kpi("⚡ En.Misurata",f"{em:,.3f} MWh","AEX","blue"),unsafe_allow_html=True)
        with c2: st.markdown(kpi("📡 En.Maggiorata",f"{emg:,.3f} MWh","Perdite rete","blue"),unsafe_allow_html=True)
        with c3: st.markdown(kpi("💼 Fat.Lordo",f"€ {fl:,.2f}","Pre-fee","gold"),unsafe_allow_html=True)
        with c4: st.markdown(kpi("💰 Fat.Netto",f"€ {fn:,.2f}","Totale fattura","green"),unsafe_allow_html=True)
        with c5: st.markdown(kpi("📈 EUR/MWh",f"€ {eur:,.2f}","Prezzo netto","gold"),unsafe_allow_html=True)
        st.divider()

        # Confronto mesi — mese selezionato evidenziato
        st.markdown('<div class="section-title">Confronto con gli altri mesi</div>',unsafe_allow_html=True)
        dall=df_fd[df_fd['Impianto']==imp].copy()
        dall['MWh']=dall['En_Mis_kWh']/1000
        dall['Mese']=pd.Categorical(dall['Mese'],categories=MESI,ordered=True); dall=dall.sort_values('Mese')
        c1,c2=st.columns(2)
        with c1:
            cb=[imp_color if m==mese else "#dde3ef" for m in dall['Mese']]
            fig=go.Figure(go.Bar(x=dall['Mese'],y=dall['MWh'],marker_color=cb,
                text=[f"{v:,.1f}" for v in dall['MWh']],textposition='inside',textfont_color='white'))
            fig.update_layout(title=f"MWh — {mese} evidenziato",plot_bgcolor=PBG,paper_bgcolor=PBG,
                font_color=FC,height=300,xaxis=dict(gridcolor=GC),yaxis=dict(gridcolor=GC),margin=dict(t=40,b=20))
            st.plotly_chart(fig,use_container_width=True)
        with c2:
            cb2=[imp_color if m==mese else "#dde3ef" for m in dall['Mese']]
            fig2=go.Figure(go.Bar(x=dall['Mese'],y=dall['Fat_Netto'],marker_color=cb2,
                text=[f"€{v:,.0f}" for v in dall['Fat_Netto']],textposition='inside',textfont_color='white'))
            fig2.update_layout(title=f"Fat.Netto — {mese} evidenziato",plot_bgcolor=PBG,paper_bgcolor=PBG,
                font_color=FC,height=300,xaxis=dict(gridcolor=GC),yaxis=dict(gridcolor=GC),margin=dict(t=40,b=20))
            st.plotly_chart(fig2,use_container_width=True)

        if not df_ap.empty:
            apm=df_ap[(df_ap['Impianto']==imp)&(df_ap['Mese']==mese)]
            if not apm.empty and apm.iloc[0]['E_Teorica']>0:
                st.markdown('<div class="section-title">Analisi Perdite</div>',unsafe_allow_html=True)
                ar=apm.iloc[0]
                c1,c2,c3=st.columns(3)
                with c1: st.metric("E.Teorica",f"{ar['E_Teorica']:,.1f} MWh")
                with c2: st.metric("E.Reale",f"{ar['E_Reale']:,.1f} MWh")
                with c3:
                    dp=ar['Delta_pct']*100 if abs(ar['Delta_pct'])<=1 else ar['Delta_pct']
                    s="✅ OK" if abs(dp)<=8 else ("⚡ Medio" if abs(dp)<=20 else "⚠ Alto")
                    st.metric("ΔE%",f"{dp:+.1f}%",delta=s)

        st.divider()
        de3=pd.DataFrame([{'Impianto':imp,'Mese':mese,'En.Mis(MWh)':em,'En.Magg(MWh)':emg,
            'Fat.Lordo(€)':fl,'Fee(€)':fee,'Fat.Netto(€)':fn,'€/MWh':eur}])
        dl_excel({f'{imp[:20]}_{mese}':de3},f"{imp.replace(' ','_')}_{mese}.xlsx",f"📥 Esporta {mese} Excel")
    else:
        st.info(f"Nessun dato per {imp} — {mese}.")

st.divider()
st.markdown("<center><small style='color:#8a9ab5'>Portfolio AM Dashboard &nbsp;|&nbsp; Q1 2026</small></center>",unsafe_allow_html=True)
